import openpyxl
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox
from tkinter.ttk import Combobox  # Import Combobox

def merge_similar_cells(filepath, column):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    prev_value = None
    start_row = None

    max_row = ws.max_row
    for row in range(1, max_row + 1):
        current_value = ws[f"{column}{row}"].value

        if current_value == prev_value:
            if not start_row:
                start_row = row - 1
        else:
            if start_row:
                cell_range = f"{column}{start_row}:{column}{row-1}"
                ws.merge_cells(cell_range)
                start_row = None

        prev_value = current_value

    # Handling the end case
    if start_row:
        cell_range = f"{column}{start_row}:{column}{max_row}"
        ws.merge_cells(cell_range)

    wb.save(filepath)


def select_column():
    """Function to show a dropdown for column selection and return the selected column."""
    win = tk.Toplevel()
    win.title("Select Column")

    label = tk.Label(win, text="Select a column:")
    label.pack(padx=10, pady=10)

    columns = [chr(i) for i in range(65, 91)]  # A to Z
    combo = Combobox(win, values=columns)
    combo.pack(padx=10, pady=10)
    combo.set("A")

    def on_ok():
        win.selected_column = combo.get()
        win.destroy()

    btn_ok = tk.Button(win, text="OK", command=on_ok)
    btn_ok.pack(padx=10, pady=20)

    win.wait_window()  # Wait until the window is closed
    return win.selected_column

def main():
    root = tk.Tk()
    root.withdraw()

    filepath = filedialog.askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx")])
    if not filepath:
        return

    column = select_column()  # Use the new function to get column selection

    if column:
        merge_similar_cells(filepath, column.upper())
        messagebox.showinfo("Success", "Cells merged successfully!")

if __name__ == "__main__":
    main()